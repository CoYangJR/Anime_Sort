#!/usr/bin/python
# -*- coding: utf-8 -*-
import os
from openpyxl import load_workbook
import shutil
import time
from datetime import datetime
from openpyxl.styles import Font, colors, Alignment
import json
import requests

database_path=r'C:\Users\yjr\source\repos\Anime_Sort\Anime_Sort'
url=r'https://api.bgm.tv/subject/'

def load_excel(path):
    global wb,ws_Local_Floder,ws_Database,ws_Search_Path
    if (os.path.isfile(os.path.join(path,r'Anime_list.xlsx'))):
        wb = load_workbook('Anime_list.xlsx')
        ws_Local_Floder=wb[r'Local Folder']
        ws_Database=wb[r'Database']
        ws_Search_Path=wb[r'Search Path']
    else:
        print('Anime list is not found!')
    return

def get_search_path(ws):#获取查询路径
    search_list=[]
    for row in ws.iter_rows(min_row=2, max_col=1, max_row=ws.max_row):
        for cell in row:
            search_list.append(cell.value)
    return search_list

def get_existed_file(ws):#获取已查找文件
    file_set=set()
    for i in range(1,len(ws['B'])):
        file_set.add(ws['B'][i].value)
    return file_set

def lwalk(top, max_level=10000):#遍历文件目录
    if max_level==0:
        return
    dirs, nondirs = [], []
    with os.scandir(top) as it:
        for entry in it:
            if entry.is_dir():
                dirs.append(entry.name)
            else:
                nondirs.append(entry.name)
        yield top, dirs, nondirs
        for dirname in dirs:
            new_path = os.path.join(top, dirname)
            yield from lwalk(new_path, max_level-1)
    return
def get_new_file(file_dir):#查询路径
    file_list=[]
    #exclude = set([r'Chronicle',r'CDs',r'Scans',r'SPs',r'Interviews',r'NCOP&NCED',r'Previews','BD Illustration Gallery',r'Bonus',r'Extras',r'fonts',r'Fonts',r'Galleries',r'BDMV',r'CERTIFICATE'])
    for root, dirs,files in lwalk(file_dir,1):
        for file in dirs:
            path=os.path.join(root,file)
            file_list.append(path)
    return file_list

def save_excel():
    if (not os.path.exists(os.path.join(database_path,r'backup'))):
        os.mkdir(os.path.join(database_path,r'backup'))
    new_name=r'Backup_'+time.strftime('%Y-%m-%d_%H%M%S',time.localtime(time.time()))+r'.xlsx'
    shutil.copyfile(os.path.join(database_path,r'Anime_list.xlsx'),os.path.join(database_path,r'backup',new_name))
    wb.save(r'Anime_list.xlsx')

def get_date():
    now = datetime.now()
    year = str(int(now.strftime('%Y')))
    month = str(int(now.strftime('%m')))
    day = str(int(now.strftime('%d')))
    return year+'/'+month+'/'+day

def refresh_list(ws_lf,ws_sp):
    search_list=get_search_path(ws_sp)#获取搜索目录
    file_set=get_existed_file(ws_lf)#获取已搜索的文件
    file_list=[]
    for search_path in search_list:#获取搜索目录里的文件
        file_list+=get_new_file(search_path)
    new_file=[]
    for i in file_list:#去除已搜索
        if (os.path.basename(i) not in file_set):
            new_file.append(i)
    st=ws_lf.max_row+1
    for i in new_file:#更新Local Folder表
        print(os.path.basename(i))
        ws_lf.append([get_date(),os.path.basename(i),None,0,i])
    ed=ws_lf.max_row
    for i in range(st,ed+1):#更改格式
            ws_lf['A'+str(i)].alignment = Alignment(horizontal='center', vertical='center')
            ws_lf['B'+str(i)].alignment = Alignment(horizontal='right', vertical='center')
            ws_lf['C'+str(i)].alignment = Alignment(horizontal='left', vertical='center')
            ws_lf['D'+str(i)].alignment = Alignment(horizontal='center', vertical='center')
            ws_lf['E'+str(i)].alignment = Alignment(horizontal='left', vertical='center')
    save_excel()
    return

def upgrade_information(ws_db,path,bgm_id):
    new_row=ws_db.max_row+1
    rs = json.loads(requests.get(url+str(bgm_id)).text)
    print('get_requests')
    ws_db['A'+str(new_row)]=new_row-2
    ws_db['B'+str(new_row)]=new_row-2
    ws_db['C'+str(new_row)]=bgm_id
    ws_db['D'+str(new_row)]=os.path.basename(path)
    ws_db['E'+str(new_row)]=rs['name_cn']
    ws_db['F'+str(new_row)]=rs['name']
    ws_db['J'+str(new_row)]=rs['summary']
    if 'eps' in rs.keys():
        ws_db['K'+str(new_row)]=rs['eps']
    ws_db['L'+str(new_row)]=rs['air_date']
    ws_db['M'+str(new_row)]=rs['images']['large']
    return

def fill_database(ws_lf,ws_db):
    for i in range(2,ws_lf.max_row):
        if (ws_lf['D'+str(i)].value==0):
            ws_lf['D'+str(i)]=1
            print(time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(time.time())))
            print(os.path.basename(ws_lf['E'+str(i)].value))
            upgrade_information(ws_db,ws_lf['E'+str(i)].value,ws_lf['C'+str(i)].value)
            print('upgraded')
            save_excel()

if __name__ == '__main__':
    load_excel(database_path)
    #refresh_list(ws_Local_Floder,ws_Search_Path)
    fill_database(ws_Local_Floder,ws_Database)
    #upgrade_information(ws_Database,r'\\nas\D\Anime\VCB\[VCB-Studio] Bungo Stray Dogs Dead Apple [Ma10p_1080p]',209596)
    print('-----END-----')